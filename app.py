# -*- coding: utf-8 -*-
"""
Application Flask pour la gestion des fiches de biodiversité marine
"""
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_from_directory  # type: ignore
from openpyxl import load_workbook  # type: ignore
import os
import secrets

app = Flask(__name__)
# Génère une clé secrète aléatoire ou utilise une variable d'environnement
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(16))

# Chemin vers le fichier Excel
EXCEL_FILE = 'donnees.xlsx'
IMAGE_FOLDER = 'images'

def lire_donnees():
    """Lit les données du fichier Excel et retourne une liste de dictionnaires"""
    if not os.path.exists(EXCEL_FILE):
        return []
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    donnees = []
    headers = [cell.value for cell in ws[1]]  # Première ligne = en-têtes
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Si l'ID existe
            espece = {
                'id': row[0],
                'nom': row[1],
                'categorie': row[2],
                'espece': row[3],
                'image': row[4],
                'description': row[5] if len(row) > 5 else ''
            }
            donnees.append(espece)
    
    return donnees

def sauvegarder_donnees(donnees):
    """Sauvegarde les données dans le fichier Excel"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Efface les données existantes (sauf en-têtes)
    ws.delete_rows(2, ws.max_row)
    
    # Écrit les nouvelles données
    for idx, espece in enumerate(donnees, start=2):
        ws.cell(row=idx, column=1, value=espece['id'])
        ws.cell(row=idx, column=2, value=espece['nom'])
        ws.cell(row=idx, column=3, value=espece['categorie'])
        ws.cell(row=idx, column=4, value=espece['espece'])
        ws.cell(row=idx, column=5, value=espece['image'])
        ws.cell(row=idx, column=6, value=espece['description'])
    
    wb.save(EXCEL_FILE)

def obtenir_categories():
    """Retourne la liste unique des catégories"""
    donnees = lire_donnees()
    categories = list(set([d['categorie'] for d in donnees if d['categorie']]))
    return sorted(categories)

def generer_nouvel_id():
    """Génère un nouvel ID unique pour une espèce"""
    donnees = lire_donnees()
    if not donnees:
        return 1
    max_id = max([d['id'] for d in donnees])
    return max_id + 1

@app.route('/')
def index():
    """Page d'accueil avec les catégories"""
    categories = obtenir_categories()
    return render_template('index.html', categories=categories)

@app.route('/categorie/<nom_categorie>')
def categorie(nom_categorie):
    """Liste des espèces d'une catégorie"""
    donnees = lire_donnees()
    especes = [d for d in donnees if d['categorie'] == nom_categorie]
    return render_template('especes.html', categorie=nom_categorie, especes=especes)

@app.route('/fiche/<int:espece_id>')
def fiche(espece_id):
    """Fiche détaillée d'une espèce"""
    donnees = lire_donnees()
    espece = next((d for d in donnees if d['id'] == espece_id), None)
    
    if not espece:
        return "Espèce non trouvée", 404
    
    return render_template('fiche.html', espece=espece)

@app.route('/ajouter', methods=['GET', 'POST'])
def ajouter():
    """Ajoute une nouvelle espèce"""
    if request.method == 'POST':
        nouveau_nom = request.form.get('nom', '')
        nouvelle_categorie = request.form.get('categorie', '')
        nouvelle_espece = request.form.get('espece', '')
        nouvelle_image = request.form.get('image', '')
        nouvelle_description = request.form.get('description', '')
        
        # Crée la nouvelle espèce
        nouvelle_fiche = {
            'id': generer_nouvel_id(),
            'nom': nouveau_nom,
            'categorie': nouvelle_categorie,
            'espece': nouvelle_espece,
            'image': nouvelle_image,
            'description': nouvelle_description
        }
        
        # Ajoute aux données et sauvegarde
        donnees = lire_donnees()
        donnees.append(nouvelle_fiche)
        sauvegarder_donnees(donnees)
        
        return redirect(url_for('fiche', espece_id=nouvelle_fiche['id']))
    
    # GET: affiche le formulaire
    categories = obtenir_categories()
    return render_template('ajouter.html', categories=categories)

@app.route('/modifier/<int:espece_id>', methods=['POST'])
def modifier(espece_id):
    """Modifie les informations d'une espèce"""
    nouveau_nom = request.form.get('nom', '')
    nouvelle_espece = request.form.get('espece', '')
    nouvelle_description = request.form.get('description', '')
    
    donnees = lire_donnees()
    for espece in donnees:
        if espece['id'] == espece_id:
            espece['nom'] = nouveau_nom
            espece['espece'] = nouvelle_espece
            espece['description'] = nouvelle_description
            break
    
    sauvegarder_donnees(donnees)
    return redirect(url_for('fiche', espece_id=espece_id))

@app.route('/images/<filename>')
def serve_image(filename):
    """Sert les images depuis le dossier images/"""
    return send_from_directory(IMAGE_FOLDER, filename)

@app.route('/api/especes')
def api_especes():
    """API JSON pour obtenir toutes les espèces"""
    donnees = lire_donnees()
    return jsonify(donnees)

if __name__ == '__main__':
    # Vérifie que le fichier Excel existe
    if not os.path.exists(EXCEL_FILE):
        print(f"ATTENTION: Le fichier {EXCEL_FILE} n'existe pas!")
        print("Créez un fichier Excel avec les colonnes: ID, nom, catégorie, especes, image, description")
    
    # Lance le serveur Flask
    # Mode développement local
    app.run(host='0.0.0.0', port=5000, debug=True)
