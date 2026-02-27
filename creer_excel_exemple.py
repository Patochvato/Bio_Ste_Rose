# -*- coding: utf-8 -*-
"""
Script pour créer un fichier Excel d'exemple pour l'application Bio Sainte-Rose
Exécutez ce script si vous n'avez pas encore de fichier donnees.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

def creer_excel_exemple():
    """Crée un fichier Excel d'exemple avec la structure correcte"""
    
    # Créer un nouveau classeur
    wb = Workbook()
    ws = wb.active
    ws.title = "Espèces Marines"
    
    # En-têtes
    headers = ['ID', 'nom', 'catégorie', 'especes', 'image', 'description']
    
    # Style des en-têtes
    header_fill = PatternFill(start_color="0077BE", end_color="0077BE", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Écrire les en-têtes
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Données d'exemple
    exemples = [
        {
            'id': 1,
            'nom': 'Poisson-perroquet à bosse',
            'categorie': 'Poissons',
            'espece': 'Bolbometopon muricatum',
            'image': 'perroquet.jpg',
            'description': 'Grand poisson herbivore reconnaissable à sa bosse frontale prononcée. Joue un rôle important dans la formation du sable corallien.'
        },
        {
            'id': 2,
            'nom': 'Tortue verte',
            'categorie': 'Reptiles',
            'espece': 'Chelonia mydas',
            'image': 'tortue.jpg',
            'description': 'Tortue marine herbivore qui se nourrit principalement d\'herbiers marins et d\'algues.'
        },
        {
            'id': 3,
            'nom': 'Corail cerveau',
            'categorie': 'Coraux',
            'espece': 'Diploria labyrinthiformis',
            'image': 'corail_cerveau.jpg',
            'description': 'Corail massif dont la surface rappelle les circonvolutions du cerveau humain.'
        },
        {
            'id': 4,
            'nom': 'Éponge tonneau',
            'categorie': 'Éponges',
            'espece': 'Xestospongia muta',
            'image': 'eponge.jpg',
            'description': 'Grande éponge en forme de tonneau, peut vivre plusieurs siècles.'
        },
        {
            'id': 5,
            'nom': 'Poisson-ange français',
            'categorie': 'Poissons',
            'espece': 'Pomacanthus paru',
            'image': 'ange_francais.jpg',
            'description': 'Poisson coloré avec un corps noir et des écailles bordées de jaune doré.'
        },
        {
            'id': 6,
            'nom': 'Langouste royale',
            'categorie': 'Crustacés',
            'espece': 'Panulirus argus',
            'image': 'langouste.jpg',
            'description': 'Crustacé nocturne reconnaissable à ses longues antennes et sa carapace épineuse.'
        },
        {
            'id': 7,
            'nom': 'Raie aigle',
            'categorie': 'Raies',
            'espece': 'Aetobatus narinari',
            'image': 'raie_aigle.jpg',
            'description': 'Raie élégante avec des taches blanches sur fond sombre, nage avec grâce.'
        },
        {
            'id': 8,
            'nom': 'Murène verte',
            'categorie': 'Poissons',
            'espece': 'Gymnothorax funebris',
            'image': 'murene.jpg',
            'description': 'Poisson serpentiforme qui vit dans les crevasses des récifs coralliens.'
        }
    ]
    
    # Écrire les données
    for row_idx, exemple in enumerate(exemples, start=2):
        ws.cell(row=row_idx, column=1, value=exemple['id'])
        ws.cell(row=row_idx, column=2, value=exemple['nom'])
        ws.cell(row=row_idx, column=3, value=exemple['categorie'])
        ws.cell(row=row_idx, column=4, value=exemple['espece'])
        ws.cell(row=row_idx, column=5, value=exemple['image'])
        ws.cell(row=row_idx, column=6, value=exemple['description'])
    
    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 60
    
    # Sauvegarder
    filename = 'donnees.xlsx'
    wb.save(filename)
    print(f"✅ Fichier '{filename}' créé avec succès!")
    print(f"\nContenu : {len(exemples)} espèces d'exemple")
    print("\n📝 Remarque : Les noms d'images sont fictifs.")
    print("   Remplacez-les par les noms réels de vos photos dans le dossier 'images/'")
    print("\n🎯 Vous pouvez maintenant modifier ce fichier avec Excel ou LibreOffice")

if __name__ == '__main__':
    if os.path.exists('donnees.xlsx'):
        reponse = input("⚠️  Le fichier 'donnees.xlsx' existe déjà. Voulez-vous le remplacer? (oui/non): ")
        if reponse.lower() not in ['oui', 'o', 'yes', 'y']:
            print("Opération annulée.")
            exit()
    
    try:
        creer_excel_exemple()
    except Exception as e:
        print(f"❌ Erreur lors de la création du fichier: {e}")
